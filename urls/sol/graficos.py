from statistics import median

from flask import Blueprint, render_template, request, send_file
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import column

from database.setor_dao import SetorDAO

bp_graficos = Blueprint('graficos', __name__, url_prefix='/sol/graficos')


@bp_graficos.route('/menu')
def menu():
    return render_template('sol/graficos/menu.html')


def _dados_organograma():
    """Busca e estrutura os dados para o ECharts."""
    dao_setor = SetorDAO()

    setores = dao_setor.read_all()
    setores_nodes = []

    for setor in setores:
        servicos_nodes = []
        for servico in setor.tb_servico_collection:
            servicos_nodes.append({
                'name': servico.nme_servico + ' R$ ' + str(servico.vlr_servico),
            })
        setores_nodes.append({
            'name': setor.nme_setor + "-" + setor.sgl_setor,
            'children': servicos_nodes
        })

    return {
        'name': 'Todos os Setores',
        'children': setores_nodes
    }


@bp_graficos.route('/hierarquia')
def hierarquia():
    return render_template('sol/graficos/hierarquia.html', dados=_dados_organograma())


@bp_graficos.route('/exportar_excel')
def exportar_excel():
    # Lendo todos os setores ativos
    dao_setor = SetorDAO()
    setores = dao_setor.read_by_filters([('sts_setor', '=', 'A')])

    # 1. Cria o Workbook do openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Carta de Serviços"

    # --- INÍCIO DA FORMATAÇÃO DO CABEÇALHO ---
    # 1.1. Definição dos Estilos
    header_font = Font(bold=True, color='FFFFFF')  # Fonte em negrito e branca
    # Preenchimento em uma cor que contraste (ex: um azul escuro)
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    # Borda fina para todas as laterais
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # 1.2. Definição dos Cabeçalhos
    headers = {
        'A1': 'Setores',
        'B1': 'Serviços',
        'C1': 'Dias Necessários',
        'D1': 'Valor do Serviço'
    }
    # 1.3. Aplicação dos Estilos nos Cabeçalhos
    for cell_ref, value in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.border = thin_border
        cell.fill = header_fill

    # --- FIM DA FORMATAÇÃO DO CABEÇALHO ---

    # 2. Adiciona as linhas de exemplo
    linha = 2
    num_servico = 0
    soma = 0
    maior = 0
    menor = 99999999999999
    for setor in setores:
        for servico in setor.tb_servico_collection:
            if servico.sts_servico == 'A':
                ws.cell(row=linha, column=1, value=f'{setor.nme_setor} ({setor.sgl_setor})')
                ws.cell(row=linha, column=2, value=f'{servico.nme_servico}')
                ws.cell(row=linha, column=3, value=servico.num_dias_servico)
                ws.cell(row=linha, column=4, value=f'R$ {servico.vlr_servico}')
                num_servico += 1
                soma += servico.vlr_servico
                if maior < servico.vlr_servico:
                    maior = servico.vlr_servico
                if menor > servico.vlr_servico:
                    menor = servico.vlr_servico
                linha += 1
        linha += 1
    media = soma / num_servico
    ws.cell(row=linha, column=3, value='Minimo')
    ws.cell(row=linha, column=4, value=f'R$ {menor}')
    linha += 1
    ws.cell(row=linha, column=3, value='Maximo')
    ws.cell(row=linha, column=4, value=f'R$ {maior}')
    linha += 1
    ws.cell(row=linha, column=3, value='Médio')
    ws.cell(row=linha, column=4, value=f'R$ {media}')

    # Opcional: Ajusta a largura das colunas para melhor visualização
    # (Requer importar 'get_column_letter' de openpyxl.utils se precisar de ajustes mais finos)
    for coluna in ws.columns:
        max_length = 0
        column = coluna[0].column_letter  # Obtém o cabeçalho da coluna (ex: 'A')
        for cell in coluna:
            try:  # Tenta encontrar o comprimento máximo do conteúdo
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 3. Salva o Workbook em um objeto BytesIO (em memória)
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)  # Retorna ao início do stream para leitura

    # 4. Envia o arquivo como resposta
    nome_arquivo = "carta_servico.xlsx"

    return send_file(
        excel_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_arquivo
    )
